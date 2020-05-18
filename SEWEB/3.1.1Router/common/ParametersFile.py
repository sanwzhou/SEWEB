#coding=gbk
import os
import sys
import unittest
from openpyxl import Workbook
from openpyxl import load_workbook
from common.ReadConfig import getpath,getParameter
from common.LogGen import LogGen
logger = LogGen(Logger = 'ParametersFile').getlog()
tmppath = os.path.dirname(os.path.abspath('.')) + '/tmp/'


def GetParameters(excel = getpath('parameterpath1')):
    u'''���� �����ļ�'''
    # excel = getpath('parameterpath')
    excel2 = tmppath + getpath('parameterExcel')

    ProductmodelP = getParameter('ProductmodelP')
    natsessionP = getParameter('natsessionP')
    memoryP = getParameter('memoryP')
    wifiload2Gp = getParameter('wifiload2Gp')
    wifiload5Gp = getParameter('wifiload5Gp')
    sessionLifep = getParameter('sessionLifep')
    passwdErrNumP = getParameter('passwdErrNumP')
    loginSpanP = getParameter('loginSpanP')
    AppUpgradeP = getParameter('AppUpgradeP')
    licenseP = getParameter('licenseP')
    portTateWANp = getParameter('portTateWANp')
    portTateLANp = getParameter('portTateLANp')
    staticRp = getParameter('staticRp')
    policyP = getParameter('policyP')
    NetSniperP = getParameter('NetSniperP')
    staticMapP = getParameter('staticMapP')
    natRuleP = getParameter('natRuleP')
    bandingVPNp = getParameter('bandingVPNp')
    FirewallP = getParameter('FirewallP')
    addGroupP = getParameter('addGroupP')
    timePlanP = getParameter('timePlanP')
    actionMp = getParameter('actionMp')
    QQnumP = getParameter('QQnumP')
    alinumP = getParameter('alinumP')
    DomainFilerP = getParameter('DomainFilerP')
    FlowruleP = getParameter('FlowruleP')
    vlanPortP = getParameter('vlanPortP')
    portVlanP = getParameter('portVlanP')
    portmirrorP = getParameter('portmirrorP')
    L2tpP = getParameter('L2tpP')
    pptpP = getParameter('pptpP')
    ipsecP = getParameter('ipsecP')
    ibindingPp = getParameter('ibindingPp')
    vpnNumP = getParameter('vpnNumP')
    userNumP = getParameter('userNumP')
    ipmac = getParameter('ipmac')
    pppoeSp = getParameter('pppoeSp')
    webAutnp = getParameter('webAutnp')
    remoteAuthP = getParameter('remoteAuthP')
    BlacklistP = getParameter('BlacklistP')
    netShareP = getParameter('netShareP')
    FtpP = getParameter('FtpP')
    APnumP = getParameter('APnumP')
    RoamP = getParameter('RoamP')
    ssidnumP = getParameter('ssidnumP')
    hwNatP = getParameter('hwNatP')
    sfnatP = getParameter('sfnatP')
    line = [ProductmodelP,natsessionP, memoryP,wifiload2Gp,wifiload5Gp,sessionLifep,passwdErrNumP,loginSpanP,AppUpgradeP,
            licenseP,portTateWANp,portTateLANp,staticRp,policyP,NetSniperP,staticMapP,natRuleP,bandingVPNp,FirewallP,
            addGroupP,timePlanP,actionMp,QQnumP,alinumP,DomainFilerP,FlowruleP,vlanPortP,portVlanP,portmirrorP,L2tpP,
            pptpP,ipsecP,ibindingPp,hwNatP,sfnatP
        ,vpnNumP,userNumP,ipmac,pppoeSp,webAutnp,remoteAuthP,BlacklistP,netShareP,FtpP,APnumP,RoamP,ssidnumP]

    wbb = Workbook()
    wsb = wbb.active
    wbb.save(excel2)

    wb = load_workbook(excel)
    sheets = wb.get_sheet_names()
    first = sheets[0]
    ws = wb.get_sheet_by_name(first)

    wb2 = load_workbook(excel2)
    sheets2 = wb2.get_sheet_names()
    first2 = sheets2[0]
    ws2 = wb2.get_sheet_by_name(first2)

    location = 'C%s'
    y = []
    x = 1
    while x < 300: #������300��
        # print(ws[location % x].value)
        if ws[location % x].value != None:
            y.append(ws[location % x].value)
        x += 1
    # print(y)
    i = 0
    n = 1
    while i < (len(y)):
        # print(y[i])
        for linex in line:
            if linex in y[i]:
                # print(i,i + 1)
                p = ('D%s' % (i + 1))
                # print(p)
                value0 = str(ws[p].value)
                # print(value0)
                if '֧��' in value0:
                    if '��' in value0:
                        valuex = str(value0).split(r'��')[1].split(r'��')[0]
                    elif '(' in value0:
                        valuex = str(value0).split(r'(')[1].split(r')')[0]
                    # print('1', valuex)
                elif 'DDR3 ' in value0:
                    valuex = str(value0).split(r' ')[1]
                    # print('2', valuex)
                else:
                    valuex = value0
                    # print('3', value0)

                ws2['A%s' % n] = y[i]
                ws2['B%s' % n] = valuex
                # print(valuex)
                wb2.save(excel2)
                n += 1
        i += 1
    # print(p,ws[p].value)
    logger.info(u'�������ļ��ѻ�ȡ������')
    logger.info(u'GetParameters over')

def delParametersFile():
    u'''ɾ�����ɵ� �����ļ�'''
    parameterExcel = getpath('parameterExcel')
    file = parameterExcel
    path_xlsx= tmppath  # �����ļ���ţ����أ�·��
    sys.path.append(r'%s' % path_xlsx)
    files_cvs = os.listdir(r'%s' % path_xlsx)  # os.listdir(path) ����pathָ�����ļ��а������ļ����ļ��е����ֵ��б�
    for filename_cvs in files_cvs:
        portion_cvs = os.path.splitext(filename_cvs)  # splitext()���ڷ��� �ļ�������չ�� Ԫ��
        # print(portion_cvs)
        if file.split(r'.')[0] == portion_cvs[0]:  # ����ļ�������"�������20"
            if portion_cvs[1] == '.xlsx':  # ��׺�� .csv
                # ��������ļ����ͺ�׺����ֱ���޸����ƵĻ� ������ļ�������һ��Ŀ¼��
                filenamedir_cvs = (r'%s' % path_xlsx) + filename_cvs
                os.remove(filenamedir_cvs)
    logger.info(u'�������ļ���ɾ��')
    logger.info(u'delParametersFile over')


# GetParameters()