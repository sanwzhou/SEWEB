#coding=gbk
#@Time: 2019/3/19 0019 18:52
#@swzhou
'''
从性能指标中得到对应参数值
'''
import os
from openpyxl import load_workbook
from common.ReadConfig import getpath,getParameter
tmppath = os.path.dirname(os.path.abspath('.')) + '/tmp/'

def getExcelValue(parameter):
    excel = tmppath + getpath('parameterExcel')
    wb = load_workbook(excel)
    sheets = wb.get_sheet_names()
    sheets_first = sheets[0]
    ws = wb.get_sheet_by_name(sheets_first)

    location = 'A%s'
    y = []
    x = 1
    while x < 300:  # 假设有300行
        # print(ws[location % x].value)
        if ws[location % x].value != None:
            if str(ws[location % x].value).isspace():#判断末尾是否是空格
                y.append(u'%s' % str(ws[location % x].value).split(r' ')[0])
            else:
                y.append(u'%s' % str(ws[location % x].value))
            # y.append(str(ws[location % x].value))
        x += 1
    # print(y)

    i = 0
    # print(len(y))
    while i < (len(y)):
        # print('"%s"' % y[i],i)
        if str(parameter) == str(y[i]):
            # print(parameter,str(y[i]))
            parameter1 = ('A%s' % (i))
            # print(parameter1)
            value = ('B%s' % (i+1))
            # print(parameter,ws[value].value)
            # print(str(ws[value].value).isspace())
            # if str(ws[value].value).isspace():  # 判断末尾是否是空格
            #     return str(ws[value].value).split(r' ')[0]
            # else:
            #     return ws[value].value
            if parameter == '产品型号': #产品型号 名称一般有空格
                return str(ws[value].value)
            elif str(' ') in str(ws[value].value)[0]:  # 判断前是否是空格
                return str(ws[value].value).split(r' ')[1]
            else:
                return str(ws[value].value).split(r' ')[0]
        i += 1

# portTateWANp = getParameter('ProductmodelP')
# print(portTateWANp)
# print(getExcelValue(parameter=portTateWANp))
# NetSniperP = getParameter('APnumP')
# print(getExcelValue(parameter=NetSniperP))


